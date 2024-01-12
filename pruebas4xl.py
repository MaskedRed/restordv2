import tkinter as tk
import openpyxl as xl
from openpyxl.styles import Font, Border, Side
from datetime import datetime
import os

#CHECKMARK UBER O GRATIS, INTRODUCIR FECHAS POR ORDEN

app = tk.Tk()
app.title("Burritos_Norteños")

class Burritos:
    def __init__(self, master):
        self.master = master
        self.master.title("Órdenes Restaurante")
        #Titulo
        self.nombree = tk.Label(self.master, text="Bienvenido al robot Burrito3000", font=('Helvetiva', 25))
        self.nombree.pack()
        #Frame general
        self.framegen = tk.Frame(self.master)
        self.framegen.pack(side=tk.TOP)
        # Frame botones
        self.botframe = tk.Frame(self.framegen)
        self.botframe.pack(anchor=tk.NW, side=tk.LEFT)
            #Frame botones burritos
        #self.burrframe = tk.Frame(self.botframe)
        #self.burrframe.grid(column=0, row=1)
            # Frame botones Bebida
        #self.bebframe = tk.Frame(self.botframe)
       # self.bebframe.grid(column=0, row=2)

        #Frame para frame de ord
        self.frameord = tk.Frame(self.framegen)
        self.frameord.pack(padx=10, pady=10, side=tk.LEFT)
        #Frame para subframes ordenes
        self.orders_frame = tk.Frame(self.frameord)
        self.orders_frame.grid(padx=10, pady=10, column=0, row=0)
        self.no_orders = tk.Label(self.orders_frame, text='No hay órdenes activas por el momento', padx=10,
                             font=('Helvetica', 18))

        #Diccionario de precios
        self.prices = {}
        #Diccionario de conteo
        self.count = {}
        #Diccionario para modificaciones
        self.mod_count = {}

        #Lista de tipos de articulos
        self.listatype = []



        # Suma del dia
        self.totaldia = 0

        # Número de orden
        self.numpedido = 0

        #Lista de pedidos generales, contiene: al principio cantidad de burritos, numero de pedido y total
        self.lista = []


        #Aquí se agregan las label para manipularlas (para que se actualize el contador en pantalla)
        self.productos = []


        #Label Burritos
        #self.labelburros = tk.Label(self.burrframe, text="Burritos", font=('Helvetiva', 20))
        #Row burros
        self.mainrow = 0

        #Introducir productos aqui (self.input_prod(Nombre, precio, tipo) tipo debe ser consistente
        self.input_prod('Asada', 35, 'Burritos')
        self.input_prod('Deshebrada', 35, 'Burritos')
        self.input_prod('Cochinita', 35, 'Burritos')
        self.input_prod('Adobada', 35, 'Burritos')
        self.input_prod('Frijol con Queso', 35, 'Burritos')

        self.input_prod('Agua de Vainilla', 27, 'Bebidas')
        self.input_prod('Refill', 7, 'Bebidas')
        self.input_prod('Boing', 16, 'Bebidas')
        self.input_prod('Refresco', 16, 'Bebidas')

        self.input_prod('Postre del día', 20, 'Postres')
        self.input_prod('Donas', 10, 'Postres')


        #Función para poner en display "No ordenes activas ahora"
        self.no_ord()

        self.retrieve_info()

    def retrieve_info(self):
        current_date = datetime.now().strftime("%Y-%m-%d")
        month_year_folder = datetime.now().strftime("%Y-%m")
        filename = f"Report_{current_date}.xlsx"
        file_path = os.path.join(month_year_folder, filename)
        if os.path.exists(file_path):
            workbook = xl.load_workbook(file_path)
            sheetlista = workbook['Sheet2']
            for row in sheetlista.iter_rows(values_only=True):
                self.lista.append(row)
            self.numpedido = self.lista[-1][-2]

    #Para introducir productos (burritos)
    def input_prod(self, nombre, precio, tipo):
        #Pone nombre con precio en el diccionario
        self.prices[nombre] = precio
        #Agrega al contador y lo pone en cero
        self.count[nombre] = 0
        #" " " modificado " " " " "
        self.mod_count[nombre] = 0
        if self.listatype:
            for type in self.listatype:
                if tipo in type:
                    type[0] += 1
                    self.productos.append(self.hacerbotones(nombre=nombre, frame=type[1], row=type[0]))
                    type.append(nombre)
                    break
            else:
                freimm = tk.Frame(self.botframe)
                freimm.grid(column=0, row=self.mainrow)
                label = tk.Label(freimm, text=f'{tipo}', font=('Helvetiva', 18))
                label.grid(row=0, sticky=tk.W)
                self.mainrow += 1
                cosolista = [1, freimm, tipo]
                self.productos.append(self.hacerbotones(nombre=nombre, frame=freimm, row=1))
                self.listatype.append(cosolista)
                for type in self.listatype:
                    if tipo in type:
                        type.append(nombre)


        else:
            freimm = tk.Frame(self.botframe)
            freimm.grid(column=0, row=self.mainrow)
            label = tk.Label(freimm, text=f'{tipo}', font=('Helvetiva', 18))
            label.grid(row=0, sticky=tk.W)
            self.mainrow += 1
            cosolista = [1, freimm, tipo]
            self.productos.append(self.hacerbotones(nombre=nombre, frame=freimm, row=1))
            self.listatype.append(cosolista)
            for type in self.listatype:
                if tipo in type:
                    type.append(nombre)


    #Función para agregar burritos
    def ag_burros(self, nombre):
        self.count[nombre] += 1
        return self.count[nombre]

    #" " quitar "
    def qu_burros(self, nombre):
        self.count[nombre] -= 1
        return self.count[nombre]

    #Para regresar el valor en el contador de burritos
    def not_burros(self, nombre):
        return self.count[nombre]

    #Creador del texto para display en ordenes
    def texto(self, dict):
        lines = []
        only_type = [types[2:] for types in self.listatype]
        for things in only_type:
            stuff = sum(dict.get(item, 0) for item in things)
            if stuff != 0:
                lines.append(f'-{things[0]}')
                lines.extend(
                    [f'{item}: {count}' if count != 0 else '' for item, count in dict.items() if item in things])
                lines.append('--------')  # Add a blank line after the 'Burritos' section

        return '\n'.join(line for line in lines if line).strip()


    #Función que calcula y regresa el total del pedido
    def ordenar(self, dict):
        total = 0
        for item, price in self.prices.items():
            quantity = dict.get(item, 0)  # Get the quantity or default to 0 if not found
            total += price * quantity

        return total


    #Crear botones
    def hacerbotones(self, nombre, frame, row):
        #Hace una label para los botones
        label = tk.Label(frame, text=f'{nombre}: {self.count[nombre]}', font=('Helvetica', 15))
        label.grid(row=row, column=0, padx=5)

        #Botones con comandos
        botonmas = tk.Button(frame, text='+', command=lambda: self.actualizar(label, self.ag_burros, nombre),
                             font=('Helvetica', 15))
        botonmas.grid(row=row, column=1, padx=8)

        botonmenos = tk.Button(frame, text='-', command=lambda: self.actualizar(label, self.qu_burros, nombre),
                               font=('Helvetica', 15))
        botonmenos.grid(row=row, column=2, padx=8)

        return label

    #Función que actualiza la label de los botones
    def actualizar(self, label, comando, nombre):
        # Update the label's text using the provided command
        nuevo_contador = comando(nombre)
        label.config(text=f'{nombre}: {nuevo_contador}')

    #Muestra si no hay ordenes disponibles
    def no_ord(self):
        children = self.orders_frame.winfo_children()
        #print("Number of children:", len(children))

        if len(children) == 1:
            #print("Displaying self.no_orders")
            self.no_orders.grid()
        else:
            #print("Forgetting self.no_orders")
            self.no_orders.grid_forget()

    #Función para tomar ordenes
    def tomar_orden(self):
        #Cantidad a cobrar
        pedido = self.ordenar(self.count)
        if pedido != 0:
            # Establecer número de pedido
            self.numpedido += 1

            #Orden en texto
            text_output = self.texto(self.count)

            #Agregar a la lista de pedidos
            self.lista.append([self.count[counter] for counter in self.count.keys()] + [self.numpedido] + [pedido])

            #Contadores volviendo a 0
            self.clear()

            #Crea el frame donde se muestra la orden activa
            self.create_frame_order(order_text=f'Orden#{self.numpedido}', detail=text_output, row=self.numpedido)

            do_excel()

    #Crear frame
    def create_frame_order(self, order_text, detail, row):
        #Frame donde cabe lo de abajo
        order_frame = tk.Frame(self.orders_frame, bd=2, relief=tk.GROOVE)
        order_frame.grid(row=row)

        #Frame para mostrar num y boyones
        in_order_frame = tk.Frame(order_frame)
        in_order_frame.pack(padx=15)

        order_label = tk.Label(in_order_frame, text=order_text, padx=10, font=('Helvetica', 15), background='CadetBlue')
        order_label.pack(side=tk.LEFT)


        edit_button = tk.Button(in_order_frame, text="Edit",
                                command=lambda: self.mod_orden(order_frame, order_text), font=('Helvetica', 15))
        edit_button.pack(side=tk.LEFT, padx=5)


        delete_button = tk.Button(in_order_frame, text="Delete",
                                  command=lambda: self.borrar(order_frame, order_text), font=('Helvetica', 15))
        delete_button.pack(side=tk.LEFT, padx=5)


        finish_button = tk.Button(in_order_frame, text="Finalizar",
                                  command=lambda: self.fin_orden(order_frame), font=('Helvetica', 15))
        finish_button.pack(side=tk.LEFT, padx=5)


        #Frame para display de los detalles
        detail_frame = tk.Frame(order_frame)
        detail_frame.pack(padx=8, pady=8)

        detail_label = tk.Label(detail_frame, text=detail, padx=10, font=('Helvetica', 15))
        detail_label.pack(side=tk.LEFT)


        self.no_ord()


#Funciones especiales (lo que más me tomó tiempo)

    #Función para modificar órdenes
    def mod_orden(self, order_frame, order_text):
        #Obtener número de pedido de la label
        num_ped = int(order_text.strip('Orden#')) - 1

        #Esto es para debugging
        #print(num_ped)
        #print(self.lista[num_ped])
        #print(self.count)

        #Iterar en el diccionario de conteo para obtener los nombres y valores actuales
        for i, (name, value) in enumerate(self.count.items()):
            #Sumarlo a la lista de pedido para obtener valor modificado
            self.mod_count[name] = int(value) + int(self.lista[num_ped][i])

        #Cantidad nueva
        pedido = self.ordenar(self.mod_count)

        #Agregar nueva version a la lista de pedidos (sustituyendo la vieja)
        self.lista[num_ped] = [self.mod_count[counter] for counter in self.mod_count.keys()] + [num_ped + 1] + [pedido]

        #Texto nuevo
        text_output = self.texto(self.mod_count)

        #Modificar los detalles entrando a la label
        detail_label = order_frame.winfo_children()[1].winfo_children()[0]
        detail_label.config(text=text_output)

        #Regresar el contador de modificaciones a cero
        for name in self.mod_count.keys():
            self.mod_count[name] = 0

        self.clear()


    #Limpiar valores
    def clear(self):
        #Regresa el contador del diccionario a cero
        for key in self.count.keys():
            self.count[key] = 0
        #Actualiza la label de los botones iterando de la label de productos y el nombre sacado del contador
        for prod, key in zip(self.productos, self.count.keys()):
            self.actualizar(prod, self.not_burros, key)


    #Finalizar orden (solo quita la orden de la pantalla)
    def fin_orden(self, order_frame):
        order_frame.destroy()
        self.no_ord()


    #Elimina la orden
    def borrar(self, order_frame, order_text):

        num_ped = int(order_text.strip('Orden#')) - 1

        #Cambiar a solo ceros
        self.lista[num_ped] = [0]*len(self.count) + [num_ped + 1] + [0]

        order_frame.destroy()

        self.no_ord()

        self.clear()


    def cierre(self): #Debe tener suma de self.ordenes[3], crear la listbox abajo
        for ped in self.lista:
            self.totaldia += ped[-1]

        do_excel()
        self.clear()


burr = Burritos(app)
total_dia = 0



#Frame ordenar o limpiar
burriros = tk.Frame(burr.botframe)
burriros.grid(column=0, row=3)


    #Ordenar
boton_ordenar = tk.Button(burriros, text='Agregar orden', command=burr.tomar_orden, font=('Helvetiva', 15))
boton_ordenar.pack(side=tk.LEFT, padx=6, pady=6)

    #Limpiar valores
boton_limpiar = tk.Button(burriros, text='Limpiar', command=burr.clear, font=('Helvetiva', 15))
boton_limpiar.pack(side=tk.RIGHT, padx=5, pady=5)

#Cerrar el dia
boton_ordenar = tk.Button(burriros, text='Cerrar el día', command=burr.cierre, font=('Helvetiva', 15))
boton_ordenar.pack(side=tk.LEFT, padx=6, pady=6)



# Excel shit

archivo = xl.Workbook()
sheet = archivo.active
sheetlista = archivo.create_sheet('Sheet2')

def do_excel():

    sheet.cell(1, 1).value = '#'
    i = 0
    j = 0

    for i, ord in enumerate(burr.lista):
        sheet.cell(i + 2, 1).value = ord[-2]

    modified_data = [sublist[:-2] + sublist[-1:] for sublist in burr.lista]

    for i, prod in enumerate(burr.count.keys()):
        sheet.cell(1, i + 2).value = prod
        i += 1
    sheet.cell(1, i + 2).value = 'Total'

    for i, sublist in enumerate(modified_data):
        j = 0
        for subsub in sublist:
            sheet.cell(i + 2, j + 2).value = subsub
            j += 1

    for k, prod in enumerate(burr.count.keys()):
        totali = sum(int(sub[k]) for sub in modified_data)
        total_total = sum(int(sub[-1]) for sub in modified_data)
        sheet.cell(i + 3, 2 + k).value = totali
        sheet.cell(i + 3, 2 + len(burr.count.keys())).value = total_total
        sheet.cell(i+3, 1).value = 'Total'

    for n, cosos in enumerate(burr.lista):
        for m, coso in enumerate(cosos):
            sheetlista.cell(1+n, 1+m).value = coso





    #Formatting

    #Main Table formatting
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))

    for row in sheet.iter_rows(min_row=1, max_row=i+3, min_col=1, max_col=j+1):
        for cell in row:
            cell.border = border

    #Adjusting cell size
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    #Bolding the top and debolding the table
    table_font = Font(bold=False)
    header_font = Font(bold=True)

    for cell in sheet[1]:
        cell.font = header_font

    for colu in sheet.iter_rows(min_row=i+3, max_row=i+4, min_col=1, max_col=j+1):
        for cell in colu:
            cell.font = header_font

    for row in sheet.iter_rows(min_row=1, max_row=i+2, min_col=1, max_col=j+1):
        for cell in row:
            cell.font = table_font


    folder_date = datetime.now()
    current_date = datetime.now().strftime("%Y-%m-%d")
    month_year_folder = folder_date.strftime("%Y-%m")

    if not os.path.exists(month_year_folder):
        os.makedirs(month_year_folder)

    filename = os.path.join(month_year_folder, f"{current_date}.xlsx")
    archivo.save(filename)




app.mainloop()
