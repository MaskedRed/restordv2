import tkinter as tk
import openpyxl as xl
from openpyxl.styles import Font, Border, Side
from datetime import datetime
import os, sys
from openpyxl.utils import get_column_letter

# CHECKMARK UBER O GRATIS, INTRODUCIR FECHAS POR ORDEN

app = tk.Tk()
app.title("Burritos_Norteños")

font_stuff_20 = ('Helvetica', 10)
font_stuff_25 = ('Helvetica', 14)
font_stuff_30 = ('Helvetica', 15)
font_bold_25 = 'Helvetica 13 bold'

def combo_1(dictio, lista):
    burritos = dictio['Burrito']
    refrescos = dictio['Refresco']
    refill = dictio['Refill']
    dict_new = dictio.copy()


    if dict_new['Finalizado'] == 0:
        if burritos >= 2 and refrescos >= 1:
            dict_new['Burrito'] -= 2
            dict_new['Refresco'] -= 1
            dict_new['Combo1'] += 1
            dict_new['Ahorro'] += 11
            lista.append(dict_new)

        elif burritos >= 2 and refill >= 1:
            dict_new['Burrito'] -= 2
            dict_new['Refill'] -= 1
            dict_new['Combo1'] += 1
            dict_new['Ahorro'] += 2
            lista.append(dict_new)

        else:
            dict_new['Finalizado'] = 1
            lista.append(dict_new)
            print('Hola')

def combo_2(dictio, lista):
    burritos = dictio['Burrito']
    dict_new = dictio.copy()

    if dict_new['Finalizado'] == 0:
        if burritos >= 4:
            dict_new['Burrito'] -= 4
            dict_new['Combo2'] += 1
            dict_new['Ahorro'] += 15
            lista.append(dict_new)
        else:
            dict_new['Finalizado'] = 1
            lista.append(dict_new)

def combo_3(dictio, lista):
    burritos = dictio['Burrito']
    dict_new = dictio.copy()

    if dict_new['Finalizado'] == 0:
        if burritos >= 10:
            dict_new['Burrito'] -= 10
            dict_new['Combo3'] += 1
            dict_new['Ahorro'] += 45
            lista.append(dict_new)
        else:
            dict_new['Finalizado'] = 1
            lista.append(dict_new)



class Burritos:
    def __init__(self, master):
        self.master = master
        self.master.title("Órdenes Restaurante")
        # Titulo
        self.nombree = tk.Label(self.master, text="Burritos Norteños Punto de Venta", font='Helvetica 20 bold')
        self.nombree.pack()
        # Frame general
        self.framegen = tk.Frame(self.master)
        self.framegen.pack(side=tk.TOP)
        # Frame botones
        self.botframe = tk.Frame(self.framegen)
        self.botframe.pack(anchor=tk.NW, side=tk.LEFT)

        # Frame para frame de ord
        self.frameord = tk.Frame(self.framegen)
        self.frameord.pack(padx=10, pady=10, side=tk.LEFT)
        # Frame para subframes ordenes
        self.orders_frame = tk.Frame(self.frameord)
        self.orders_frame.grid(padx=10, pady=10, column=0, row=0)
        self.no_orders = tk.Label(self.orders_frame, text='No hay órdenes activas por el momento', padx=10,
                                  font=font_stuff_25)
        # Frame botones abajo
        self.burriros = tk.Frame(self.botframe)
        self.burriros.grid(column=0, row=3)



        # Boolean var para uber
        self.uber_var = tk.BooleanVar()
        self.uber_checkbutton = tk.Checkbutton(self.burriros, text="Uber", variable=self.uber_var,
                                               font=font_stuff_25)
        self.uber_checkbutton.pack(padx=10)

        # Boolean for rappi
        self.rappi_var = tk.BooleanVar()
        self.rappi_checkbutton = tk.Checkbutton(self.burriros, text="Rappi", variable=self.rappi_var,
                                                font=font_stuff_25)
        self.rappi_checkbutton.pack(padx=10)

        # Boolean for gratis
        self.gratis_var = tk.BooleanVar()
        self.gratis_checkbutton = tk.Checkbutton(self.burriros, text="Gratis", variable=self.gratis_var,
                                                 font=font_stuff_25)
        self.gratis_checkbutton.pack(padx=10)



        # Diccionario de precios
        self.prices = {}
        # Diccionario de conteo
        self.count = {}
        # Diccionario para modificaciones
        self.mod_count = {}

        # Combos
        self.n_combo_1 = 0
        self.n_combo_2 = 0
        self.n_combo_3 = 0


        # Lista de tipos de articulos
        self.listatype = []


        # Lista combos
        self.lista_combos = [combo_3, combo_2, combo_1]
        # Lista combinaciones combos
        self.lista_totales = []


        # Suma del dia
        self.totaldia = 0

        # Número de orden
        self.numpedido = 0

        # Lista de pedidos generales, contiene: al principio cantidad de burritos, numero de pedido y total
        self.lista = []

        # Aquí se agregan las label para manipularlas (para que se actualize el contador en pantalla)
        self.productos = []

        self.mainrow = 0




        # Introducir productos aqui (self.input_prod(Nombre, precio, tipo) tipo debe ser consistente
        self.input_prod('Asada', 45, 'Burritos')
        self.input_prod('Deshebrada', 45, 'Burritos')
        self.input_prod('Cochinita', 45, 'Burritos')
        self.input_prod('Adobada', 45, 'Burritos')
        self.input_prod('Frijol con Queso', 45, 'Burritos')

        self.input_prod('Agua de Vainilla', 27, 'Bebidas')
        self.input_prod('Refill', 7, 'Bebidas')
        self.input_prod('Boing', 16, 'Bebidas')
        self.input_prod('Refresco', 16, 'Bebidas')

        self.input_prod('Postre $10', 10, 'Postres')
        self.input_prod('Postre $15', 15, 'Postres')
        self.input_prod('Postre $20', 20, 'Postres')
        self.input_prod('Postre $25', 25, 'Postres')
        self.input_prod('Postre $30', 30, 'Postres')

        # Función para poner en display "No ordenes activas ahora"
        self.no_ord()

        self.retrieve_info()

    def retrieve_info(self):
        if getattr(sys, 'frozen', False):
            base_path = os.path.dirname(sys.argv[0])
        else:
            base_path = os.path.dirname(os.path.abspath(__file__))

        current_date = datetime.now().strftime("%Y-%m-%d")
        month_year_folder = datetime.now().strftime("%Y-%m")
        filename = f"Report_{current_date}.xlsx"

        file_path = os.path.join(base_path, month_year_folder, filename)

        if os.path.exists(file_path):
            workbook = xl.load_workbook(file_path)
            sheetlista = workbook['Sheet2']
            for row in sheetlista.iter_rows(values_only=True):
                self.lista.append(row)
            self.numpedido = self.lista[-1][-2]

    # Para introducir productos (burritos)
    def input_prod(self, nombre, precio, tipo):
        # Pone nombre con precio en el diccionario
        self.prices[nombre] = precio
        # Agrega al contador y lo pone en cero
        self.count[nombre] = 0
        # " " " modificado " " " " "
        self.mod_count[nombre] = 0
        if self.listatype:
            for type in self.listatype:
                if tipo in type:
                    type[0] += 1
                    self.productos.append(self.hacerbotones(nombre=nombre, frame=type[1], row=type[0]))
                    type.append(nombre)
                    break
            else:
                freimm = tk.Frame(self.botframe)
                freimm.grid(column=0, row=self.mainrow)
                label = tk.Label(freimm, text=f'{tipo}', font=font_stuff_30, fg='Red')
                label.grid(row=0, sticky=tk.W)
                self.mainrow += 1
                cosolista = [1, freimm, tipo]
                self.productos.append(self.hacerbotones(nombre=nombre, frame=freimm, row=1))
                self.listatype.append(cosolista)
                for type in self.listatype:
                    if tipo in type:
                        type.append(nombre)


        else:
            freimm = tk.Frame(self.botframe)
            freimm.grid(column=0, row=self.mainrow)
            label = tk.Label(freimm, text=f'{tipo}', font=font_stuff_30, fg='Red')
            label.grid(row=0, sticky=tk.W)
            self.mainrow += 1
            cosolista = [1, freimm, tipo]
            self.productos.append(self.hacerbotones(nombre=nombre, frame=freimm, row=1))
            self.listatype.append(cosolista)
            for type in self.listatype:
                if tipo in type:
                    type.append(nombre)

    # Función para agregar burritos
    def ag_burros(self, nombre):
        self.count[nombre] += 1
        return self.count[nombre]

    # " " quitar "
    def qu_burros(self, nombre):
        self.count[nombre] -= 1
        return self.count[nombre]

    # Para regresar el valor en el contador de burritos
    def not_burros(self, nombre):
        return self.count[nombre]

    # Creador del texto para display en ordenes
    def texto(self, dict):
        lines = []
        gratis = 1 if self.gratis_var.get() else 0
        totali = 0
        if gratis == 0:
            totali = self.ordenar(dict)
        elif gratis == 1:
            totali = 0

        only_type = [types[2:] for types in self.listatype]
        combo_1_text = f'Paquete Uno: {self.n_combo_1}' if self.n_combo_1 != 0 else ''
        combo_2_text = f'Paquete Dos: {self.n_combo_2}' if self.n_combo_2 != 0 else ''
        combo_3_text = f'Paquete Tres: {self.n_combo_3}' if self.n_combo_3 != 0 else ''
        lines.append(combo_1_text)
        lines.append(combo_2_text)
        lines.append(combo_3_text)

        if combo_1_text != '' or combo_2_text != '' or combo_3_text != '':
            lines.append('_____')

        for things in only_type:
            stuff = sum(dict.get(item, 0) for item in things)
            if stuff != 0:
                lines.append(f'-{things[0]}')
                lines.extend(
                    [f'{item}: {count}' if count != 0 else '' for item, count in dict.items() if item in things])
                lines.append('__')  # Add a blank line after the 'Burritos' section
        lines.append(f'Total: {totali}')

        return '\n'.join(line for line in lines if line).strip()

    # Función que calcula y regresa el total del pedido
    def ordenar(self, dict):
        total = 0
        burritos = 0
        refrescos = 0
        refill = 0
        total_menos = 0

        self.lista_totales.clear()
        self.n_combo_3 = 0
        self.n_combo_2 = 0
        self.n_combo_1 = 0

        for item, quantity in dict.items():
            if item in self.listatype[0]:
                burritos += quantity
            if item == 'Refresco' or item == 'Boing':
                refrescos += quantity
            if item == 'Refill':
                refill += quantity

        dict_total = {
            'Burrito': 0,
            'Refresco': 0,
            'Refill': 0,
            'Combo1': 0,
            'Combo2': 0,
            'Combo3': 0,
            'Ahorro': 0,
            'Finalizado': 0
        }
        dict_total['Burrito'] = burritos
        dict_total['Refresco'] = refrescos
        dict_total['Refill'] = refill
        self.lista_totales.append(dict_total)

        while True:
            all_finalizado = all(totales['Finalizado'] == 1 for totales in self.lista_totales)
            if all_finalizado:
                break

            for totales in self.lista_totales.copy():
                for combo in self.lista_combos:
                    combo(totales, self.lista_totales)
                self.lista_totales.remove(totales)

        new_list_totales = [totales['Ahorro'] for totales in self.lista_totales]
        index_lis = new_list_totales.index(max(new_list_totales))

        # Seleccionado el maximo
        dict_seleccionado = self.lista_totales[index_lis]
        print(dict_seleccionado)

        ahorro = dict_seleccionado['Ahorro']
        self.n_combo_1 = dict_seleccionado['Combo1']
        self.n_combo_2 = dict_seleccionado['Combo2']
        self.n_combo_3 = dict_seleccionado['Combo3']


        for item, price in self.prices.items():
            quantity = dict.get(item, 0)  # Get the quantity or default to 0 if not found
            total += price * quantity


        print(f'Uno: {self.n_combo_1}, Dos: {self.n_combo_2}, Tres: {self.n_combo_3}')

        return total - ahorro

    # Crear botones
    def hacerbotones(self, nombre, frame, row):
        # Hace una label para los botones
        label = tk.Label(frame, text=f'{nombre}: {self.count[nombre]}', font=font_bold_25)
        label.grid(row=row, column=0, padx=5)

        # Botones con comandos
        botonmas = tk.Button(frame, text='+', command=lambda: self.actualizar(label, self.ag_burros, nombre),
                             font=font_bold_25)
        botonmas.grid(row=row, column=1, padx=8)

        botonmenos = tk.Button(frame, text='-', command=lambda: self.actualizar(label, self.qu_burros, nombre),
                               font=font_bold_25)
        botonmenos.grid(row=row, column=2, padx=8)

        return label

    # Función que actualiza la label de los botones
    def actualizar(self, label, comando, nombre):
        # Update the label's text using the provided command
        nuevo_contador = comando(nombre)
        label.config(text=f'{nombre}: {nuevo_contador}')

    # Muestra si no hay ordenes disponibles
    def no_ord(self):
        children = self.orders_frame.winfo_children()
        # print("Number of children:", len(children))

        if len(children) == 1:
            # print("Displaying self.no_orders")
            self.no_orders.grid()
        else:
            # print("Forgetting self.no_orders")
            self.no_orders.grid_forget()

    # Función para tomar ordenes
    def tomar_orden(self):
        stuff = 0
        # Cantidad a cobrar
        pedido = self.ordenar(self.count)
        uber = 1 if self.uber_var.get() else 0
        rappi = 1 if self.rappi_var.get() else 0
        gratis = 1 if self.gratis_var.get() else 0
        if pedido != 0:
            # Establecer número de pedido
            self.numpedido += 1

            if gratis == 0:
                stuff = pedido
            elif gratis == 1:
                stuff = 0

            paquete1 = self.n_combo_1
            paquete2 = self.n_combo_2
            paquete3 = self.n_combo_3

            text_output = self.texto(self.count)

            # Agregar a la lista de pedidos
            self.lista.append(
                [self.count[counter] for counter in self.count.keys()] + [paquete1] +
                [paquete2] + [paquete3] + [uber] + [rappi] + [gratis] + [self.numpedido] + [stuff])

            # Contadores volviendo a 0
            self.clear()

            # Crea el frame donde se muestra la orden activa
            self.create_frame_order(order_text=f'Orden#{self.numpedido}', detail=text_output, row=self.numpedido,
                                    uber=self.uber_var.get(), rappi=self.rappi_var.get(), gratis=self.gratis_var.get())

            self.gratis_var.set(False)
            self.rappi_var.set(False)
            self.uber_var.set(False)
            do_excel()

    # Crear frame
    def create_frame_order(self, order_text, detail, row, uber=False, rappi=False, gratis=False):
        # Frame donde cabe lo de abajo
        order_frame = tk.Frame(self.orders_frame, bd=2, relief=tk.GROOVE)
        order_frame.grid(row=row)

        # Frame para mostrar num y boyones
        in_order_frame = tk.Frame(order_frame)
        in_order_frame.pack(padx=15)

        order_label = tk.Label(in_order_frame, text=order_text, padx=10, font=font_stuff_25, background='CadetBlue')
        order_label.pack(side=tk.LEFT)

        if uber:
            uber_order_label = tk.Label(in_order_frame, text='(Uber)', padx=10, font=font_stuff_20,
                                        background='CadetBlue')
            uber_order_label.pack(side=tk.LEFT)

        if rappi:
            rappi_order_label = tk.Label(in_order_frame, text='(rappi)', padx=10, font=font_stuff_20,
                                         background='CadetBlue')
            rappi_order_label.pack(side=tk.LEFT)

        if gratis:
            gratis_order_label = tk.Label(in_order_frame, text='(Gratis)', padx=10, font=font_stuff_20,
                                          background='CadetBlue')
            gratis_order_label.pack(side=tk.LEFT)

        edit_button = tk.Button(in_order_frame, text="Edit",
                                command=lambda: self.mod_orden(order_frame, order_text), font=font_stuff_25)
        edit_button.pack(side=tk.LEFT, padx=5)

        delete_button = tk.Button(in_order_frame, text="Delete",
                                  command=lambda: self.borrar(order_frame, order_text), font=font_stuff_25)
        delete_button.pack(side=tk.LEFT, padx=5)

        finish_button = tk.Button(in_order_frame, text="Finalizar",
                                  command=lambda: self.fin_orden(order_frame), font=font_stuff_25)
        finish_button.pack(side=tk.LEFT, padx=5)

        # Frame para display de los detalles
        detail_frame = tk.Frame(order_frame)
        detail_frame.pack(padx=8, pady=8)

        detail_label = tk.Label(detail_frame, text=detail, padx=10, font=font_stuff_25)
        detail_label.pack(side=tk.LEFT)

        self.no_ord()

    # Funciones especiales (lo que más me tomó tiempo)

    # Función para modificar órdenes
    def mod_orden(self, order_frame, order_text):
        # Obtener número de pedido de la label
        num_ped = int(order_text.strip('Orden#')) - 1

        # Iterar en el diccionario de conteo para obtener los nombres y valores actuales
        for i, (name, value) in enumerate(self.count.items()):
            # Sumarlo a la lista de pedido para obtener valor modificado
            self.mod_count[name] = int(value) + int(self.lista[num_ped][i])


        uber = 1 if self.uber_var.get() else 0
        rappi = 1 if self.rappi_var.get() else 0
        gratis = 1 if self.gratis_var.get() else 0

        paquete1 = self.n_combo_1
        paquete2 = self.n_combo_2
        paquete3 = self.n_combo_3

        pedido = self.ordenar(self.mod_count) if gratis == 0 else 0

        # Agregar nueva version a la lista de pedidos (sustituyendo la vieja)
        self.lista[num_ped] = [self.mod_count[counter] for counter in self.mod_count.keys()] + [paquete1] + \
        [paquete2] + [paquete3] + [uber] + [rappi] + [gratis] + [num_ped + 1] + [pedido]

        # Texto nuevo
        text_output = self.texto(self.mod_count)

        # Modificar los detalles entrando a la label
        detail_label = order_frame.winfo_children()[1].winfo_children()[0]
        detail_label.config(text=text_output)

        # Regresar el contador de modificaciones a cero
        for name in self.mod_count.keys():
            self.mod_count[name] = 0

        do_excel()
        self.clear()

    # Limpiar valores
    def clear(self):
        # Regresa el contador del diccionario a cero
        for key in self.count.keys():
            self.count[key] = 0
        # Actualiza la label de los botones iterando de la label de productos y el nombre sacado del contador
        for prod, key in zip(self.productos, self.count.keys()):
            self.actualizar(prod, self.not_burros, key)

    # Finalizar orden (solo quita la orden de la pantalla)
    def fin_orden(self, order_frame):
        order_frame.destroy()
        self.no_ord()

    # Elimina la orden
    def borrar(self, order_frame, order_text):

        num_ped = int(order_text.strip('Orden#')) - 1

        # Cambiar a solo ceros
        self.lista[num_ped] = [0] * (len(self.count)) + ([0] * 8)
        self.lista[num_ped][-2] = num_ped + 1

        order_frame.destroy()

        self.no_ord()
        do_excel()
        self.clear()

    def cierre(self):  # Debe tener suma de self.ordenes[3], crear la listbox abajo
        for ped in self.lista:
            self.totaldia += ped[-1]

        do_excel()
        self.clear()


burr = Burritos(app)
total_dia = 0


# Ordenar
boton_ordenar = tk.Button(burr.burriros, text='Agregar orden', command=burr.tomar_orden, font=font_stuff_25)
boton_ordenar.pack(side=tk.LEFT, padx=6, pady=6)

# Limpiar valores
boton_limpiar = tk.Button(burr.burriros, text='Limpiar', command=burr.clear, font=font_bold_25)
boton_limpiar.pack(side=tk.RIGHT, padx=5, pady=5)

# Cerrar el dia
boton_ordenar = tk.Button(burr.burriros, text='Cerrar el día', command=burr.cierre, font=font_stuff_25)
boton_ordenar.pack(side=tk.LEFT, padx=6, pady=6)

# Excel shit

archivo = xl.Workbook()
sheet = archivo.active
sheetlista = archivo.create_sheet('Sheet2')


def do_excel():
    sheet.cell(1, 1).value = '#'
    i = 0
    j = 0

    for i, ord in enumerate(burr.lista):
        sheet.cell(i + 2, 1).value = ord[-2]

    modified_data = [sublist[:-2] + sublist[-1:] for sublist in burr.lista]

    for i, prod in enumerate(burr.count.keys()):
        sheet.cell(1, i + 2).value = prod
        i += 1
    sheet.cell(1, i + 2).value = 'Paquete_1'
    sheet.cell(1, i + 3).value = 'Paquete_2'
    sheet.cell(1, i + 4).value = 'Paquete_3'
    sheet.cell(1, i + 5).value = 'Uber'
    sheet.cell(1, i + 6).value = 'Rappi'
    sheet.cell(1, i + 7).value = 'Gratis'
    sheet.cell(1, i + 8).value = 'Total'

    for i, sublist in enumerate(modified_data):
        j = 0
        for subsub in sublist:
            sheet.cell(i + 2, j + 2).value = subsub
            j += 1

    for k, prod in enumerate(burr.count.keys()):
        totali = sum(int(sub[k]) for sub in modified_data)
        total_total = sum(int(sub[-1]) for sub in modified_data)
        sheet.cell(i + 3, 2 + k).value = totali
        sheet.cell(i + 3, 3 + k).value = f'=SUM({get_column_letter(3 + k)}{i + 2}:{get_column_letter(3 + k)}{2})'
        sheet.cell(i + 3, 4 + k).value = f'=SUM({get_column_letter(4 + k)}{i + 2}:{get_column_letter(4 + k)}{2})'
        sheet.cell(i + 3, 5 + k).value = f'=SUM({get_column_letter(5 + k)}{i + 2}:{get_column_letter(5 + k)}{2})'
        sheet.cell(i + 3, 5 + len(burr.count.keys())).value = 'N/A'
        sheet.cell(i + 3, 6 + len(burr.count.keys())).value = 'N/A'
        sheet.cell(i + 3, 7 + len(burr.count.keys())).value = 'N/A'
        sheet.cell(i + 3, 8 + len(burr.count.keys())).value = total_total
        sheet.cell(i + 3, 1).value = 'Total'

    for b, tipo in enumerate(burr.listatype):
        mod_type = [element for element in burr.count.keys() if element in tipo]
        total = 0
        for col in sheet.iter_cols():
            for cell in col:
                if cell.value in mod_type:
                    total += sheet.cell(i + 3, cell.col_idx).value

        sheet.cell(i + 4, (2 * b) + 1).value = str(tipo[2])
        sheet.cell(i + 4, (2 * b) + 2).value = total

    for n, cosos in enumerate(burr.lista):
        for m, coso in enumerate(cosos):
            sheetlista.cell(1 + n, 1 + m).value = coso

    # Formatting

    # Main Table formatting
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))

    for row in sheet.iter_rows(min_row=1, max_row=i + 3, min_col=1, max_col=j + 1):
        for cell in row:
            cell.border = border

    # Adjusting cell size
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Bolding the top and debolding the table
    table_font = Font(bold=False)
    header_font = Font(bold=True)

    for colu in sheet.iter_rows(min_row=i + 3, max_row=i + 4, min_col=1, max_col=j + 1):
        for cell in colu:
            cell.font = header_font

    for row in sheet.iter_rows(min_row=1, max_row=i + 2, min_col=1, max_col=j + 1):
        for cell in row:
            cell.font = table_font

    for cell in sheet[1]:
        cell.font = header_font

    if getattr(sys, 'frozen', False):
        base_path = os.path.dirname(sys.argv[0])
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))

    folder_date = datetime.now()
    current_date = datetime.now().strftime("%Y-%m-%d")
    month_year_folder = folder_date.strftime("%Y-%m")

    full_path = os.path.join(base_path, month_year_folder)

    if not os.path.exists(full_path):
        os.makedirs(full_path)

    filename = os.path.join(full_path, f"Report_{current_date}.xlsx")
    archivo.save(filename)


app.mainloop()